import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

# Streamlit app title
st.title("Reconciliation Report Generator")

# Sidebar for selecting the report type
report_type = st.sidebar.selectbox("Select Report Type", ["GST Reconciliation", "Debit Note Reconciliation", "Combined GST Reconciliation"])

# Define color for highlighting "Mismatch" and "Debit Note"
mismatch_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
debit_note_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Define red_fill
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Define yellow_fill

# GST Reconciliation Report
if report_type == "GST Reconciliation":
    st.header("GST Reconciliation Report Generator")
    tally_file = st.file_uploader("Upload Tally Purchase Register", type=["xlsx"])
    gstr_file = st.file_uploader("Upload GSTR-2B Data", type=["xlsx"])

    def generate_gst_report(tally_file, gstr_file):
        # Read Tally Purchase Register
        tally_df = pd.read_excel(tally_file, skiprows=9, dtype=str)

        # Read GSTR-2B data
        gstr_df = pd.read_excel(gstr_file, skiprows=4, dtype=str)

        # Read GSTR-CDNR data
        gstr_cdnr_df = pd.read_excel(gstr_file, sheet_name="B2B-CDNR", skiprows=3, dtype=str)

        # Define correct column names for Tally Purchase Register
        tally_df.columns = [
            "Date", "Particulars", "Voucher_Type", "Voucher_No", "Supplier_Invoice_No",
            "Supplier_Invoice_Date", "GSTIN", "Gross_Total", "Purchase_Accounts",
            "Fixed_Assets", "Direct_Expenses", "Indirect_Expenses", "IGST", "CGST", "SGST"
        ]

        # Define correct column names for GSTR-2B
        gstr_df.columns = [
            "GSTIN", "Trade_Name", "Invoice_No", "Invoice_Type", "Invoice_Date",
            "Invoice_Value", "Place_of_Supply", "Reverse_Charge", "Taxable_Value", "Integrated_Tax",
            "Central_Tax", "State_UT_Tax", "Cess", "GSTR_IFF_Period", "GSTR_IFF_Filing_Date",
            "ITC_Availability", "Reason", "Applicable_Tax_Rate", "Source", "IRN", "IRN_Date"
        ]

        # Define correct column names for GSTR-CDNR
        gstr_cdnr_df.columns = [
            "GSTIN", "Trade_Legal_Name", "Invoice_No", "Note_Type", "Note_Supply_Type",
            "Note_Date", "Invoice_Value", "Place_of_Supply", "Supply_Attract_Reverse_Charge", "Taxable_Value",
            "Integrated_Tax", "Central_Tax", "State_UT_Tax", "Cess", "GSTR_1_IFF_GSTR_5_Period",
            "GSTR_1_IFF_GSTR_5_Filing_Date", "ITC_Availability", "Reason", "Applicable_Tax_Rate",
            "Source", "IRN", "IRN_Date"
        ]

        # Ensure 'Note_Type' exists and filter only "Debit Note" records
        if "Note_Type" in gstr_cdnr_df.columns:
            debit_note_df = gstr_cdnr_df[gstr_cdnr_df["Note_Type"].str.contains("Debit Note", case=False, na=False)]
        else:
            debit_note_df = pd.DataFrame(columns=gstr_cdnr_df.columns)

        # Remove entries where both GSTIN and Invoice No are missing
        tally_df.dropna(subset=["GSTIN", "Supplier_Invoice_No"], how="all", inplace=True)
        gstr_df.dropna(subset=["GSTIN", "Invoice_No"], how="all", inplace=True)
        debit_note_df.dropna(subset=["GSTIN", "Invoice_No"], how="all", inplace=True)

        # Convert relevant numeric columns to float
        tally_numeric_cols = ["Gross_Total", "Purchase_Accounts", "Fixed_Assets",
                              "Direct_Expenses", "Indirect_Expenses", "IGST", "CGST", "SGST"]
        gstr_numeric_cols = ["Invoice_Value", "Taxable_Value", "Integrated_Tax", "Central_Tax", "State_UT_Tax"]

        for col in tally_numeric_cols:
            tally_df[col] = pd.to_numeric(tally_df[col], errors="coerce").fillna(0)

        for col in gstr_numeric_cols:
            gstr_df[col] = pd.to_numeric(gstr_df[col], errors="coerce").fillna(0)
            if col in debit_note_df.columns:
                debit_note_df[col] = pd.to_numeric(debit_note_df[col], errors="coerce").fillna(0)

        # Compute total expense in Tally
        tally_df["Total_Expense"] = tally_df[["Purchase_Accounts", "Fixed_Assets", "Direct_Expenses", "Indirect_Expenses"]].sum(axis=1)

        # Merge Tally with GSTR-2B
        reconciliation_df_b2b = pd.merge(
            tally_df, gstr_df,
            left_on=["Supplier_Invoice_No", "GSTIN"],
            right_on=["Invoice_No", "GSTIN"],
            how="outer",
            suffixes=("_Tally", "_GSTR"),
            indicator=True
        )

        # Merge Tally with GSTR-CDNR (Debit Note Only)
        reconciliation_df_cdnr = pd.merge(
            tally_df, debit_note_df,
            left_on=["Supplier_Invoice_No", "GSTIN"],
            right_on=["Invoice_No", "GSTIN"],
            how="outer",
            suffixes=("_Tally", "_GSTR"),
            indicator=True
        )

        # Ensure B2B-CDNR (Debit Note) records are retained
        if "Note_Type" in reconciliation_df_cdnr.columns:
            reconciliation_df_cdnr = reconciliation_df_cdnr[reconciliation_df_cdnr["Note_Type"].str.contains("Debit Note", case=False, na=False)]

        # Define ₹2 tolerance threshold
        tolerance = 2.00

        # Identify Reconciliation Status
        def get_status(row):
            if row["_merge"] == "right_only":  # Exists only in GSTR (Missing in Tally)
                return "Missing in Tally"
            elif row["_merge"] == "left_only":  # Exists only in Tally (Missing in GSTR)
                return "Missing in GSTR"
            # Convert to numeric before comparison, handling potential errors
            gross_total = pd.to_numeric(row["Gross_Total"], errors="coerce")
            invoice_value = pd.to_numeric(row["Invoice_Value"], errors="coerce")
            total_expense = pd.to_numeric(row["Total_Expense"], errors="coerce")
            taxable_value = pd.to_numeric(row["Taxable_Value"], errors="coerce")
            igst = pd.to_numeric(row["IGST"], errors="coerce")
            integrated_tax = pd.to_numeric(row["Integrated_Tax"], errors="coerce")
            cgst = pd.to_numeric(row["CGST"], errors="coerce")
            central_tax = pd.to_numeric(row["Central_Tax"], errors="coerce")
            sgst = pd.to_numeric(row["SGST"], errors="coerce")
            state_ut_tax = pd.to_numeric(row["State_UT_Tax"], errors="coerce")

            # Perform the comparisons after conversion
            if (abs(gross_total - invoice_value) > tolerance or
                  abs(total_expense - taxable_value) > tolerance or
                  abs(igst - integrated_tax) > tolerance or
                  abs(cgst - central_tax) > tolerance or
                  abs(sgst - state_ut_tax) > tolerance):
                return "Mismatch"
            else:
                return "Matched"

        # Apply reconciliation logic
        reconciliation_df_b2b["Status"] = reconciliation_df_b2b.apply(get_status, axis=1)
        reconciliation_df_cdnr["Status"] = reconciliation_df_cdnr.apply(get_status, axis=1)

        # Drop merge indicator column
        reconciliation_df_b2b.drop(columns=["_merge"], inplace=True)
        reconciliation_df_cdnr.drop(columns=["_merge"], inplace=True)

        # Select only required columns
        output_df_b2b = reconciliation_df_b2b[[
            "GSTIN", "Supplier_Invoice_No", "Invoice_No", "Invoice_Value", "Gross_Total",
            "Taxable_Value", "Total_Expense", "IGST", "Integrated_Tax",
            "CGST", "Central_Tax", "SGST", "State_UT_Tax", "Status"
        ]]

        output_df_cdnr = reconciliation_df_cdnr[[
            "GSTIN", "Supplier_Invoice_No", "Invoice_No", "Invoice_Value", "Gross_Total",
            "Taxable_Value", "Total_Expense", "IGST", "Integrated_Tax",
            "CGST", "Central_Tax", "SGST", "State_UT_Tax", "Status"
        ]]

        # Combine both DataFrames into one for GSTR-2B + CDNR (Debit Note)
        combined_df = pd.concat([output_df_b2b, output_df_cdnr], ignore_index=True)

        # Save to Excel with formatting
        output_file = "GST_Reconciliation_Report_Combined.xlsx"
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            combined_df.to_excel(writer, sheet_name="GSTR-2B", index=False)

            # Access the workbook and the sheet
            workbook = writer.book
            sheet_b2b = workbook["GSTR-2B"]

            # Function to apply red highlight to mismatch rows and yellow for Debit Notes
            def highlight_rows(sheet, df):
                for i, row in df.iterrows():
                    if row["Status"] == "Mismatch":
                        for col in sheet.iter_cols(min_row=i+2, max_row=i+2, min_col=1, max_col=len(df.columns)):
                            col[0].fill = mismatch_fill
                    # Highlight Debit Note rows in yellow
                    if i >= len(output_df_b2b):  # These rows come from B2B-CDNR (Debit Note)
                        for col in sheet.iter_cols(min_row=i+2, max_row=i+2, min_col=1, max_col=len(df.columns)):
                            col[0].fill = debit_note_fill

            # Highlight mismatches and debit note rows
            highlight_rows(sheet_b2b, combined_df)

        return output_file

    if st.button("Generate GST Report"):
        if tally_file and gstr_file:
            output_file = generate_gst_report(tally_file, gstr_file)
            st.success("✅ GST Reconciliation Report Generated Successfully!")
            st.download_button(
                label="Download Report",
                data=open(output_file, "rb"),
                file_name=output_file,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.error("Please upload both Tally Purchase Register and GSTR-2B Data files.")

# Debit Note Reconciliation Report
elif report_type == "Debit Note Reconciliation":
    st.header("Debit Note Reconciliation Report Generator")
    debit_file = st.file_uploader("Upload Debit Note Register", type=["xlsx"])
    gstr_file = st.file_uploader("Upload GSTR-2B Data", type=["xlsx"])

    def generate_debit_note_report(debit_file, gstr_file):
        # Read Debit Note Register
        debit_df = pd.read_excel(debit_file, skiprows=9, dtype=str).iloc[:-1]

        # Read GSTR-CDNR data with correct header row
        gstr_cdnr_df = pd.read_excel(gstr_file, sheet_name="B2B-CDNR", skiprows=5, dtype=str)

        # Define correct column names for Debit Note Register
        debit_df.columns = [
            "Date", "Particulars", "Supplier_Invoice_No", "Credit Note Date", "Voucher Type", "Voucher_No",
            "Voucher Ref. No.", "Voucher Ref. Date", "GSTIN", "Gross_Total", "Purchase_Accounts",
            "Fixed_Assets", "IGST", "CGST", "SGST", "Round Off"
        ]

        # Define correct column names for GSTR-CDNR
        gstr_cdnr_df.columns = [
            "GSTIN_of_Supplier", "Trade_Legal_Name", "Invoice_Number", "Note_Type", "Note_Supply_Type",
            "Note_Date", "Invoice_Value", "Place_of_Supply", "Supply_Attract_Reverse_Charge", "Taxable_Value",
            "Integrated_Tax", "Central_Tax", "State_UT_Tax", "Cess", "GSTR_1_IFF_GSTR_5_Period",
            "GSTR_1_IFF_GSTR_5_Filing_Date", "ITC_Availability", "Reason", "Applicable_Tax_Rate",
            "Source", "IRN", "IRN_Date"
        ]

        # Filter records where Note_Type contains 'Credit Note' (case insensitive)
        credit_note_df = gstr_cdnr_df[gstr_cdnr_df["Note_Type"].str.contains("Credit Note", case=False, na=False)]

        # Convert only relevant numeric columns to float in debit_df
        debit_numeric_cols = ["Gross_Total", "Purchase_Accounts", "Fixed_Assets", "IGST", "CGST", "SGST"]
        debit_df[debit_numeric_cols] = debit_df[debit_numeric_cols].apply(pd.to_numeric, errors='coerce').fillna(0)

        # Convert only relevant numeric columns to float in gstr_cdnr_df
        gstr_numeric_cols = ["Invoice_Value", "Taxable_Value", "Integrated_Tax", "Central_Tax", "State_UT_Tax"]
        gstr_cdnr_df[gstr_numeric_cols] = gstr_cdnr_df[gstr_numeric_cols].apply(pd.to_numeric, errors='coerce').fillna(0)

        # Remove entries where Gross_Total and Purchase_Accounts are exactly 1.00
        debit_df = debit_df[~((debit_df['Gross_Total'] == 1.00) & (debit_df['Purchase_Accounts'] == 1.00))]

        # Calculate Total_Expense in Debit Note Register
        debit_df["Total_Expense"] = debit_df[['Purchase_Accounts', 'Fixed_Assets']].sum(axis=1)

        # Perform a full outer merge based on GSTIN + Invoice No.
        reconciliation_df = pd.merge(
            debit_df, gstr_cdnr_df,
            left_on=['Supplier_Invoice_No', 'GSTIN'],
            right_on=['Invoice_Number', 'GSTIN_of_Supplier'],
            how='outer',
            suffixes=('_DEBIT', '_GSTR'),
            indicator=True
        )

        # Fill NaN values with 0 for numeric columns
        comparison_cols = ["Gross_Total", "Invoice_Value", "Total_Expense", "Taxable_Value",
                           "IGST", "Integrated_Tax", "CGST", "Central_Tax", "SGST", "State_UT_Tax"]
        reconciliation_df[comparison_cols] = reconciliation_df[comparison_cols].fillna(0)

        # Define ₹2 tolerance threshold
        tolerance = 2.00

        # Identify Reconciliation Status with ₹2 tolerance
        def get_status(row):
            if row["_merge"] == "right_only":  # Exists only in GSTR (Missing in Tally)
                return "Missing in Tally"
            elif row["_merge"] == "left_only":  # Exists only in Tally (Missing in GSTR)
                return "Missing in GSTR"
            elif (abs(row["Gross_Total"] - row["Invoice_Value"]) > tolerance or
                  abs(row["Total_Expense"] - row["Taxable_Value"]) > tolerance or
                  abs(row["IGST"] - row["Integrated_Tax"]) > tolerance or
                  abs(row["CGST"] - row["Central_Tax"]) > tolerance or
                  abs(row["SGST"] - row["State_UT_Tax"]) > tolerance):
                return "Mismatch"
            else:
                return "Matched"

        reconciliation_df["Status"] = reconciliation_df.apply(get_status, axis=1)

        # Drop the merge indicator column
        reconciliation_df.drop(columns=["_merge"], inplace=True)

        # Select only required columns
        output_df = reconciliation_df[[ 
            "GSTIN", "Supplier_Invoice_No", "Invoice_Number", "Invoice_Value", "Gross_Total",
            "Taxable_Value", "Total_Expense", "IGST", "Integrated_Tax",
            "CGST", "Central_Tax", "SGST", "State_UT_Tax", "Status"
        ]]

        # Save the report
        output_file = "DebitNoteReconciliation_Report.xlsx"
        output_df.to_excel(output_file, index=False)

        # Load workbook for highlighting mismatches
        wb = load_workbook(output_file)
        ws = wb.active

        # Columns to check for mismatch highlighting
        columns_to_check = {
            "Invoice_Value": "Gross_Total",
            "Taxable_Value": "Total_Expense",
            "Integrated_Tax": "IGST",
            "Central_Tax": "CGST",
            "State_UT_Tax": "SGST"
        }
        status_col_idx = output_df.columns.get_loc("Status") + 1

        # Apply highlighting for mismatches
        for row in range(2, ws.max_row + 1):  # Skip header
            if ws.cell(row, status_col_idx).value == "Mismatch":
                for col_gstr, col_tally in columns_to_check.items():
                    gstr_col_idx = output_df.columns.get_loc(col_gstr) + 1
                    tally_col_idx = output_df.columns.get_loc(col_tally) + 1
                    ws.cell(row, gstr_col_idx).fill = red_fill
                    ws.cell(row, tally_col_idx).fill = red_fill

        # Save final file
        wb.save(output_file)

        return output_file

    if st.button("Generate Debit Note Report"):
        if debit_file and gstr_file:
            output_file = generate_debit_note_report(debit_file, gstr_file)
            st.success("✅ Debit Note Reconciliation Report Generated Successfully!")
            st.download_button(
                label="Download Report",
                data=open(output_file, "rb"),
                file_name=output_file,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.error("Please upload both Debit Note Register and GSTR-2B Data files.")

# Combined GST Reconciliation Report
elif report_type == "Combined GST Reconciliation":
    st.header("Combined GST Reconciliation Report Generator")
    tally_file = st.file_uploader("Upload Tally Purchase Register", type=["xlsx"])
    gstr_file = st.file_uploader("Upload GSTR-2B Data", type=["xlsx"])
    debit_file = st.file_uploader("Upload Debit Note Register", type=["xlsx"])

    def generate_combined_report(tally_file, gstr_file, debit_file):
        # Read Tally Purchase Register
        tally_df = pd.read_excel(tally_file, skiprows=9, dtype=str)

        # Read GSTR-2B data
        gstr_df = pd.read_excel(gstr_file, skiprows=5, dtype=str)

        # Read Debit Note Register
        debit_df = pd.read_excel(debit_file, skiprows=9, dtype=str)

        # Read GSTR-CDNR data with correct header row
        gstr_cdnr_df = pd.read_excel(gstr_file, sheet_name="B2B-CDNR", skiprows=5, dtype=str)

        # Define correct column names for Tally Purchase Register
        tally_df.columns = [
            "Date", "Particulars", "Voucher_Type", "Voucher_No", "Supplier_Invoice_No",
            "Supplier_Invoice_Date", "GSTIN", "Gross_Total", "Purchase_Accounts",
            "Fixed_Assets", "Direct_Expenses", "Indirect_Expenses", "IGST", "CGST", "SGST"
        ]

        # Define correct column names for GSTR-2B
        gstr_df.columns = [
            "GSTIN", "Trade_Name", "Invoice_No", "Invoice_Type", "Invoice_Date",
            "Invoice_Value", "Place_of_Supply", "Reverse_Charge", "Taxable_Value", "Integrated_Tax",
            "Central_Tax", "State_UT_Tax", "Cess", "GSTR_IFF_Period", "GSTR_IFF_Filing_Date",
            "ITC_Availability", "Reason", "Applicable_Tax_Rate", "Source", "IRN", "IRN_Date"
        ]

        # Define correct column names for Debit Note Register
        debit_df.columns = [
            "Date", "Particulars", "Supplier_Invoice_No", "Credit Note Date", "Voucher Type", "Voucher_No",
            "Voucher Ref. No.", "Voucher Ref. Date", "GSTIN", "Gross_Total", "Purchase_Accounts",
            "Fixed_Assets", "IGST", "CGST", "SGST", "Round Off"
        ]

        # Define correct column names for GSTR-CDNR
        gstr_cdnr_df.columns = [
            "GSTIN_of_Supplier", "Trade_Legal_Name", "Invoice_Number", "Note_Type", "Note_Supply_Type",
            "Note_Date", "Invoice_Value", "Place_of_Supply", "Supply_Attract_Reverse_Charge", "Taxable_Value",
            "Integrated_Tax", "Central_Tax", "State_UT_Tax", "Cess", "GSTR_1_IFF_GSTR_5_Period",
            "GSTR_1_IFF_GSTR_5_Filing_Date", "ITC_Availability", "Reason", "Applicable_Tax_Rate",
            "Source", "IRN", "IRN_Date"
        ]

        # Convert numeric columns to float
        numeric_cols_tally = ["Gross_Total", "Purchase_Accounts", "Fixed_Assets", "Direct_Expenses", 
                              "Indirect_Expenses", "IGST", "CGST", "SGST"]
        numeric_cols_gstr = ["Invoice_Value", "Taxable_Value", "Integrated_Tax", "Central_Tax", "State_UT_Tax"]
        numeric_cols_debit = ["Gross_Total", "IGST", "CGST", "SGST"]

        for col in numeric_cols_tally:
            tally_df[col] = pd.to_numeric(tally_df[col], errors='coerce').fillna(0)

        for col in numeric_cols_gstr:
            gstr_df[col] = pd.to_numeric(gstr_df[col], errors='coerce').fillna(0)

        for col in numeric_cols_debit:
            debit_df[col] = pd.to_numeric(debit_df[col], errors='coerce').fillna(0)

        # Aggregate Data by GSTIN & Trade Name
        tally_agg = tally_df.groupby("GSTIN").agg({
            "Particulars": lambda x: ', '.join(x.dropna().unique()),  # Concatenate unique names
            "Gross_Total": "sum",
            "IGST": "sum",
            "CGST": "sum",
            "SGST": "sum"
        }).reset_index()

        # Aggregate GSTR data by GSTIN while concatenating multiple Trade Names
        gstr_agg = gstr_df.groupby("GSTIN").agg({
            "Trade_Name": lambda x: ', '.join(x.dropna().unique()),  # Concatenate unique names
            "Invoice_Value": "sum",
            "Integrated_Tax": "sum",
            "Central_Tax": "sum",
            "State_UT_Tax": "sum"
        }).reset_index()

        # Aggregate Debit Note data by GSTIN while concatenating multiple Particulars
        debit_agg = debit_df.groupby("GSTIN").agg({
            "Particulars": lambda x: ', '.join(set(x)),  # Combine different Particulars
            "Gross_Total": "sum",
            "IGST": "sum",
            "CGST": "sum",
            "SGST": "sum",
        }).reset_index()

        # Aggregate GSTR-CDNR data by GSTIN while concatenating multiple Trade Names
        gstr_cdnr_agg = gstr_cdnr_df.groupby("GSTIN_of_Supplier").agg({
            "Trade_Legal_Name": lambda x: ', '.join(set(x)),  # Combine different Trade Names
            "Invoice_Value": "sum",
            "Integrated_Tax": "sum",
            "Central_Tax": "sum",
            "State_UT_Tax": "sum",
        }).reset_index()

        # Perform reconciliation based on GSTIN
        reconciliation_df = pd.merge(
            tally_agg, gstr_agg,
            left_on=["GSTIN"],
            right_on=["GSTIN"],
            how="outer",
            suffixes=("_Tally", "_GSTR"),
            indicator=True
        )

        # List of columns to fill NaN with 0 (excluding "Particulars" and "Trade Name")
        columns_to_fill = ["IGST", "CGST", "SGST", "Integrated_Tax", "Central_Tax", "State_UT_Tax"]

        # Fill NaN only for selected numeric columns
        reconciliation_df[columns_to_fill] = reconciliation_df[columns_to_fill].fillna(0)

        # Define ₹2 tolerance
        tolerance = 2.00

        # Calculate Differences
        reconciliation_df["Diff_IGST"] = reconciliation_df["IGST"] - reconciliation_df["Integrated_Tax"]
        reconciliation_df["Diff_CGST"] = reconciliation_df["CGST"] - reconciliation_df["Central_Tax"]
        reconciliation_df["Diff_SGST"] = reconciliation_df["SGST"] - reconciliation_df["State_UT_Tax"]

        # Determine Status
        def get_status(row):
            if row["_merge"] == "right_only":
                return "Missing in Tally"
            elif row["_merge"] == "left_only":
                return "Missing in GSTR"
            elif (abs(row["Diff_IGST"]) > tolerance or abs(row["Diff_CGST"]) > tolerance or abs(row["Diff_SGST"]) > tolerance):
                return "Mismatch"
            else:
                return "Matched"

        reconciliation_df["Remarks"] = reconciliation_df.apply(get_status, axis=1)

        # Drop unnecessary columns
        reconciliation_df.drop(columns=["_merge"], inplace=True)

        # Reorder columns to match output format
        final_dfg = reconciliation_df[
            [
                "GSTIN",
                "Particulars",
                "IGST",
                "CGST",
                "SGST",
                "Trade_Name",
                "Integrated_Tax",
                "Central_Tax",
                "State_UT_Tax",
                "Diff_IGST",
                "Diff_CGST",
                "Diff_SGST",
                "Remarks",
            ]
        ]
        # Sort entries alphabetically by Particulars (Tally) and Trade Name (GSTR)
        # Convert Particulars to string and sort alphabetically
        final_dfg["Particulars"] = final_dfg["Particulars"].astype(str)
        final_dfg = final_dfg.sort_values(by=["Particulars"], ascending=True)

        # Perform reconciliation for Debit Note Register
        reconciliation_df_debit = pd.merge(
            debit_agg,
            gstr_cdnr_agg,
            left_on="GSTIN",
            right_on="GSTIN_of_Supplier",
            how="outer",
            suffixes=("_DEBIT", "_GSTR"),
            indicator=True,
        )

        # Convert relevant columns to numeric before calculating differences
        reconciliation_df_debit["IGST"] = pd.to_numeric(
            reconciliation_df_debit["IGST"], errors="coerce"
        ).fillna(0)
        reconciliation_df_debit["Integrated_Tax"] = pd.to_numeric(
            reconciliation_df_debit["Integrated_Tax"], errors="coerce"
        ).fillna(0)
        reconciliation_df_debit["CGST"] = pd.to_numeric(
            reconciliation_df_debit["CGST"], errors="coerce"
        ).fillna(0)
        reconciliation_df_debit["Central_Tax"] = pd.to_numeric(
            reconciliation_df_debit["Central_Tax"], errors="coerce"
        ).fillna(0)
        reconciliation_df_debit["SGST"] = pd.to_numeric(
            reconciliation_df_debit["SGST"], errors="coerce"
        ).fillna(0)
        reconciliation_df_debit["State_UT_Tax"] = pd.to_numeric(
            reconciliation_df_debit["State_UT_Tax"], errors="coerce"
        ).fillna(0)

        # Calculate differences
        reconciliation_df_debit["Diff_IGST"] = (
            reconciliation_df_debit["IGST"] - reconciliation_df_debit["Integrated_Tax"]
        )
        reconciliation_df_debit["Diff_CGST"] = (
            reconciliation_df_debit["CGST"] - reconciliation_df_debit["Central_Tax"]
        )
        reconciliation_df_debit["Diff_SGST"] = (
            reconciliation_df_debit["SGST"] - reconciliation_df_debit["State_UT_Tax"]
        )

        def get_status_debit(row):
            if row["_merge"] == "right_only":
                return "Missing in Tally"
            elif row["_merge"] == "left_only":
                return "Missing in GSTR"
            elif (
                abs(row["Diff_IGST"]) > tolerance
                or abs(row["Diff_CGST"]) > tolerance
                or abs(row["Diff_SGST"]) > tolerance
            ):
                return "Mismatch"
            else:
                return "Matched"

        reconciliation_df_debit["Status"] = reconciliation_df_debit.apply(
            get_status_debit, axis=1
        )
        reconciliation_df_debit.drop(columns=["_merge"], inplace=True)

        # Reorder columns
        final_dfd = reconciliation_df_debit[
            [
                "GSTIN",
                "Particulars",
                "IGST",
                "CGST",
                "SGST",
                "Trade_Legal_Name",
                "Integrated_Tax",
                "Central_Tax",
                "State_UT_Tax",
                "Diff_IGST",
                "Diff_CGST",
                "Diff_SGST",
                "Status",
            ]
        ]

        # Convert numeric columns to float and negate debit values
        numeric_cols = [
            "IGST",
            "CGST",
            "SGST",
            "Integrated_Tax",
            "Central_Tax",
            "State_UT_Tax",
            "Diff_IGST",
            "Diff_CGST",
            "Diff_SGST",
        ]
        for col in numeric_cols:
            final_dfg[col] = pd.to_numeric(final_dfg[col], errors="coerce").fillna(0)
            final_dfd[col] = -pd.to_numeric(final_dfd[col], errors="coerce").fillna(0)

        # Ensure column names match
        column_mapping = {
            "GSTIN_of_Supplier": "GSTIN",
            "Trade_Legal_Name": "Trade_Name",
            "Invoice_Value": "Integrated_Tax",
            "Gross_Total": "IGST",
        }
        final_dfd.rename(columns=column_mapping, inplace=True)

        # Add remarks for debit notes
        final_dfd["Remarks"] = "Debit Note"

        # Combine both dataframes
        combined_df = pd.concat([final_dfg, final_dfd], ignore_index=True)

        # Sort so that debit notes appear immediately after their respective purchase entries
        combined_df = combined_df.sort_values(by=["Particulars"], ascending=[True])

        # Save to Excel
        output_file = "GST_Reconciliation_Summary.xlsx"
        combined_df.to_excel(output_file, index=False)

        # Load workbook for highlighting
        wb = load_workbook(output_file)
        ws = wb.active

        # Apply highlighting for debit notes
        status_col_idx = combined_df.columns.get_loc("Remarks") + 1
        for row in range(2, ws.max_row + 1):  # Skip header
            if ws.cell(row, status_col_idx).value == "Debit Note":
                for col in numeric_cols:
                    col_idx = combined_df.columns.get_loc(col) + 1
                    ws.cell(row, col_idx).fill = yellow_fill

        # Save final file
        wb.save(output_file)

        return output_file

    if st.button("Generate Combined Report"):
        if tally_file and gstr_file and debit_file:
            output_file = generate_combined_report(tally_file, gstr_file, debit_file)
            st.success("✅ Combined GST Reconciliation Report Generated Successfully!")
            st.download_button(
                label="Download Report",
                data=open(output_file, "rb"),
                file_name=output_file,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.error(
                "Please upload Tally Purchase Register, GSTR-2B Data, and Debit Note Register files."
            )